import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from tkinter import Tk, filedialog, Button, Label, Text, END, DISABLED, NORMAL, Checkbutton, BooleanVar, Entry
import threading
import os
import shutil

# === Fixed Excel 2 Path ===
EXCEL2_PATH = r"C:\Shareddoc\SHOPEE and LAZADA.xlsx"
SHEET_NAME = "ORDER (CALTEX)"

# === GUI Setup ===
root = Tk()
root.title("Shopee Received Payment")
root.geometry("854x480")
root.resizable(True, True)

# Widgets
selected_file_label = Label(root, text="No Excel 1 file selected.", wraplength=480)
selected_file_label.pack(pady=10)

status_text = Text(root, height=10, width=60)
status_text.pack(pady=10)
status_text.config(state=DISABLED)

use_today = BooleanVar(value=True)
Checkbutton(root, text="Use today's date", variable=use_today).pack()

date_entry = Entry(root, width=20)
date_entry.insert(0, "DD/MM/YYYY")
date_entry.pack(pady=5)

def log(msg):
    status_text.config(state=NORMAL)
    status_text.insert(END, msg + "\n")
    status_text.see(END)
    status_text.config(state=DISABLED)
    root.update()

def parse_date(date_str):
    if use_today.get():
        return datetime.today()
    try:
        return datetime.strptime(date_str, "%d/%m/%Y")
    except ValueError:
        log(f"⚠️ Invalid date format: {date_str}. Using today's date instead.")
        return datetime.today()

def select_file():
    file_path = filedialog.askopenfilename(title="Select Excel 1 (Shopee Report)", filetypes=[("Excel files", "*.xlsx *.xls")])
    if file_path:
        selected_file_label.config(text=f"Selected Excel 1:\n{file_path}")
        root.selected_file_path = file_path
        log(f"✅ Selected Excel 1: {file_path}")

def apply_format(source_cell, target_cell):
    try:
        target_cell.font = source_cell.font.copy()
        target_cell.fill = source_cell.fill.copy()
        target_cell.border = source_cell.border.copy()
        target_cell.alignment = source_cell.alignment.copy()
        target_cell.number_format = source_cell.number_format
    except Exception as e:
        log(f"⚠️ Formatting error: {e}")

def update_cells(ws, row, col_order_id, col_amount, col_date, order_id, amount, date):
    order_cell = ws.cell(row=row, column=col_order_id)
    if str(order_cell.value).strip() == order_id:
        amount_cell = ws.cell(row=row, column=col_amount)
        amount_cell.value = amount
        apply_format(amount_cell, amount_cell)

        date_cell = ws.cell(row=row, column=col_date)
        date_cell.value = date
        date_cell.number_format = "DD/MM/YYYY"
        apply_format(date_cell, date_cell)

        return True
    return False

def run_update():
    if not hasattr(root, 'selected_file_path') or not root.selected_file_path:
        log("❌ Please select Excel 1 file first!")
        return

    def task():
        # Disable buttons while processing
        for btn in root.winfo_children():
            if isinstance(btn, Button):
                btn.config(state=DISABLED)

        try:
            log("🚀 Starting update...")

            # --- Backup Excel 2 ---
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = EXCEL2_PATH.replace(".xlsx", f"_backup_{timestamp}.xlsx")
            shutil.copy2(EXCEL2_PATH, backup_path)
            log(f"💾 Backup created: {backup_path}")

            # Load Excel 1
            df1 = pd.read_excel(root.selected_file_path, sheet_name="Sheet1", usecols=["Order ID", "Amount"])
            df1['Order ID'] = df1['Order ID'].astype(str).str.strip()
            order_to_amount = dict(zip(df1['Order ID'], df1['Amount']))
            log("✅ Excel 1 loaded.")

            # Load Excel 2
            wb = load_workbook(EXCEL2_PATH)
            ws = wb[SHEET_NAME]
            log("✅ Excel 2 loaded.")

            # Get column indices in Excel 2
            header_map = {str(cell.value).strip().upper(): cell.column for cell in ws[1] if cell.value}
            col_order_id = header_map.get("ORDER ID".upper())
            col_amount = header_map.get("RECEIVED AMOUNT".upper())
            col_date = 14  # Column N

            if not col_order_id or not col_amount:
                log("❌ Missing ORDER ID or RECEIVED AMOUNT column in Excel 2.")
                return

            # Update orders
            unmatched_orders = []
            updated_count = 0
            excel2_order_ids = {str(ws.cell(row=row, column=col_order_id).value).strip() for row in range(2, ws.max_row + 1)}

            for order_id in order_to_amount:
                if order_id not in excel2_order_ids:
                    unmatched_orders.append(order_id)
                else:
                    for row in range(2, ws.max_row + 1):
                        order_cell = ws.cell(row=row, column=col_order_id)
                        if str(order_cell.value).strip() == order_id:
                            amount = order_to_amount[order_id]
                            date = parse_date(date_entry.get()) if not use_today.get() else datetime.today()
                            updated = update_cells(ws, row, col_order_id, col_amount, col_date, order_id, amount, date)
                            if updated:
                                updated_count += 1
                            break

            # Log results
            if unmatched_orders:
                log(f"❌ Unmatched Order IDs in Excel 2: {', '.join(unmatched_orders)}")
            else:
                log("🎉 All Order IDs from Excel 1 were found in Excel 2.")

            log(f"✅ Total Order IDs successfully updated: {updated_count} out of {len(order_to_amount)}")

            # Save the updated Excel 2
            wb.save(EXCEL2_PATH)
            log(f"💾 File saved: {EXCEL2_PATH}")
            os.startfile(EXCEL2_PATH)
            log("🎉 Done!")

        except Exception as e:
            log(f"❌ Error: {e}")

        finally:
            # Re-enable buttons
            for btn in root.winfo_children():
                if isinstance(btn, Button):
                    btn.config(state=NORMAL)

    threading.Thread(target=task).start()

# Buttons
Button(root, text="Select Excel 1 (Shopee Report)", command=select_file).pack(pady=5)
Button(root, text="Start Update", command=run_update).pack(pady=5)

root.mainloop()
